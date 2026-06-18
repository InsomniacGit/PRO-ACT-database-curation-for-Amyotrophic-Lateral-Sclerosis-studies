"""
Random Forest Results Parser - Log to Formatted Excel
======================================================
This script reads the plain-text log file produced by RF_ALSFRS_Intervals.py
(or any equivalent RF training script) and converts it into a structured,
colour-coded Excel workbook for easy comparison across datasets and horizons.

Each block in the log corresponds to one prediction file.  For every block
the parser extracts:
    - The source folder and file type (Fixed / Sliding)
    - MAE, RMSE and R² (mean-aggregated)
    - Row and column counts after preprocessing
    - Optionally, a conversion factor applied to MAE/RMSE when the target
      is expressed in interval counts rather than days (Death experiments)

Output Excel structure
----------------------
Row 1   : source file path (merged across all columns, frozen header).
Row 2   : column labels - dataset type then, for each horizon, MAE / RMSE /
          R² / Lines / Columns / separator.
Rows 3+ : data rows grouped by source folder, with blank separator rows
          inserted between groups and at Fixed-to-Sliding transitions.
Both header rows are frozen so they remain visible when scrolling.
All columns are auto-fitted to their widest content.

Cell colour coding
------------------
    Red    (code -1) : impossible value detected (negative MAE/RMSE or R² > 1)
    Orange (code  2) : very poor fit (R² < -1), suggesting model divergence
    Light orange (code 1) : poor fit (R² < 0), model worse than a mean predictor

Author: Bouclier Lucas
Data:   PROACT dataset (2022-07-29 release)
"""



import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os





# ------------------------------------------------------------------
# Encoding-tolerant file reader
# ------------------------------------------------------------------

def read_text_file_auto_encoding(path):
    """
    Read a text file trying several encodings in order of preference.

    Log files written on Windows may use cp1252 or latin-1 rather than
    UTF-8.  The function tries each encoding in sequence and falls back
    to UTF-8 with replacement characters as a last resort so that no
    file is silently skipped.

    Parameters
    ----------
    path : str
        Path to the text file to read.

    Returns
    -------
    str
        Full file content as a Python string.
    """
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue

    # Last resort: force UTF-8 and replace undecodable bytes
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()





# ------------------------------------------------------------------
# Metric extraction helper
# ------------------------------------------------------------------

def extract_metric(pattern, text):
    """
    Extract a metric value and its standard deviation from a log block.

    The function also assigns a quality code to each extracted value so
    that the Excel formatter can apply colour highlights:

        0  : value is within expected bounds (no highlight)
        1  : R² is negative (model performs worse than a mean baseline)
        2  : R² < -1 (model divergence - strong orange warning)
       -1  : impossible value (negative MAE/RMSE, or R² > 1 - red bug flag)

    NaN values are treated as absent and returned as empty strings.

    Parameters
    ----------
    pattern : str
        Regex pattern with two capture groups: value and std.
    text : str
        Log block to search.

    Returns
    -------
    tuple[str, int]
        (formatted_string, code) where formatted_string is "val ± std"
        or "" when the metric is absent or NaN.
    """
    match = re.search(pattern, text, re.UNICODE)
    if not match:
        return "", 0

    val, std = match.group(1), match.group(2)

    # NaN entries occur when a fold produces an undefined metric
    if val.lower() == "nan" or std.lower() == "nan":
        return "", 0

    code = 0  # default: no issue

    if "R²" in pattern:
        try:
            val_float = float(val)
            if val_float < -1:
                code = 2    # severe divergence
            elif val_float < 0:
                code = 1    # model worse than mean baseline
            elif val_float > 1:
                code = -1   # impossible value, indicates a bug
        except ValueError:
            code = -1

    if "MAE" in pattern or "RMSE" in pattern:
        if val.startswith("-"):
            code = -1       # negative error metric is impossible

    return f"{val} ± {std}", code





# ------------------------------------------------------------------
# Main parsing and Excel generation function
# ------------------------------------------------------------------

def parse_results_to_excel(txt_path, output_excel, convert_death_time):
    """
    Parse the RF training log file and write a formatted Excel workbook.

    The log file is split into per-file blocks at each 'Processing file'
    marker.  For every block the function extracts the folder name,
    file type (Fixed/Sliding), dataset shape, and the three evaluation
    metrics (MAE, RMSE, R²) from the mean-aggregated rows.

    When convert_death_time=True and the source folder name contains
    'Death', MAE and RMSE are multiplied by a conversion factor parsed
    from the folder name (e.g. 'Death_30' -> factor 30) to express
    errors in days rather than interval counts.

    The generated workbook includes:
    - A merged title row (row 1) showing the source log file path, frozen
      at the top so it stays visible when scrolling.
    - A column-label row (row 2) showing horizon labels and metric names,
      also frozen.
    - Data rows from row 3 onwards, grouped and separated as before.
    - All columns auto-fitted to the width of their widest content.

    Row grouping in the Excel output follows these spacing rules:
        - Sliding -> Fixed  : extra blank row inserted between groups
        - Fixed   -> Fixed  : extra blank row when the folder changes
        - Sliding -> Sliding : extra blank row when the folder changes

    Parameters
    ----------
    txt_path : str
        Path to the plain-text RF log file.
    output_excel : str
        Destination path for the Excel workbook.
    convert_death_time : bool
        If True, apply the folder-encoded day conversion factor to MAE
        and RMSE.
    """

    content = read_text_file_auto_encoding(txt_path)

    # Split log into one block per processed file
    blocks  = re.split(r"(?=Processing file\s*:)", content)
    records = []

    # ------------------------------------------------------------------
    # Horizon detection
    # ------------------------------------------------------------------
    # Horizon labels are inferred from the filenames encountered in the
    # log (e.g. "Fixed_3M.csv" -> "3M").  The list is built in
    # encounter order so that columns appear in the natural sequence.

    horizons_seen = []

    # ------------------------------------------------------------------
    # Block parsing loop
    # ------------------------------------------------------------------

    for block in blocks:
        if not block.strip().startswith("Processing file"):
            continue

        file_match = re.search(r"Processing file\s*:\s*(.+)", block)
        if not file_match:
            continue

        file_path = file_match.group(1).strip()
        folder    = file_path.split("/")[-2]
        filename  = file_path.split("/")[-1]

        # Fixed and Sliding files are kept in separate rows in the output
        base_type = "Fixed" if filename.startswith("Fixed") else "Sliding"
        file_type = f"{folder}_{base_type}"

        # Infer horizon label from filename stem (e.g. "Fixed_3M" -> "3M")
        stem          = os.path.splitext(filename)[0]   # drop extension
        parts         = stem.split("_")
        horizon_label = parts[-1] if len(parts) > 1 else stem
        if horizon_label not in horizons_seen:
            horizons_seen.append(horizon_label)

        # ------------------------------------------------------------------
        # Death-time conversion factor
        # ------------------------------------------------------------------
        # Survival predictions are stored in 90-day interval counts.
        # If the folder encodes a day factor (e.g. 'Death_30') and
        # convert_death_time is True, MAE and RMSE are scaled accordingly
        # so the results are expressed in days for easier interpretation.

        death_time_match = None
        if convert_death_time and "Death" in folder:
            if "Death_First_Symptoms" in folder:
                death_time_match = re.search(r"Death_First_Symptoms_(\d+)", folder)
            else:
                death_time_match = re.search(r"Death_(\d+)", folder)

        # Dataset shape (rows and columns after preprocessing)
        shape_match = re.search(r"=>\s*(\d+)\s*rows,\s*(\d+)\s*columns", block)
        if shape_match:
            n_lines, n_cols = map(int, shape_match.groups())
        else:
            n_lines = n_cols = ""

        # Default values for files skipped due to insufficient rows
        mae_val = rmse_val = r2_val = ""
        mae_code = rmse_code = r2_code = 0

        if "insufficient for 10-fold cross-validation" not in block:
            mae_val,  mae_code  = extract_metric(
                r"MAE\s*\(mean\)\s*:\s*([\d\.]+)\s*±\s*([\d\.]+)", block
            )
            rmse_val, rmse_code = extract_metric(
                r"RMSE\s*\(mean\)\s*:\s*([\d\.]+)\s*±\s*([\d\.]+)", block
            )
            r2_val,   r2_code   = extract_metric(
                r"R²\s*\(mean\)\s*:\s*([-\d\.]+)\s*±\s*([\d\.]+)", block
            )

        # Apply day conversion factor to MAE and RMSE when requested
        if death_time_match is not None:
            factor = int(death_time_match.group(1))
            if mae_val:
                v, s = (float(x) for x in mae_val.split(" ± "))
                mae_val = f"{v * factor:.2f} ± {s * factor:.2f}"
            if rmse_val:
                v, s = (float(x) for x in rmse_val.split(" ± "))
                rmse_val = f"{v * factor:.2f} ± {s * factor:.2f}"

        records.append({
            "type":      file_type,
            "base_type": base_type,
            "folder":    folder,
            "horizon":   horizon_label,
            "MAE":       mae_val,
            "RMSE":      rmse_val,
            "R2":        r2_val,
            "MAE_CODE":  mae_code,
            "RMSE_CODE": rmse_code,
            "R2_CODE":   r2_code,
            "LINES":     n_lines,
            "COLUMNS":   n_cols,
        })

    df = pd.DataFrame(records)

    # ------------------------------------------------------------------
    # Assemble final row list with blank separator rows
    # ------------------------------------------------------------------
    # Blank rows are inserted between dataset groups and at Fixed/Sliding
    # type transitions so that the Excel layout is easy to scan visually.
    # Each data row contains one cell group per horizon, ordered according
    # to horizons_seen so that columns are consistent across all rows.

    final_rows       = []
    previous_folder    = None
    previous_base_type = None

    for file_type in df["type"].unique():
        subset        = df[df["type"] == file_type]
        curr_folder    = subset.iloc[0]["folder"]
        curr_base_type = subset.iloc[0]["base_type"]

        # Build a horizon -> record mapping for this file_type
        horizon_map = {r["horizon"]: r for _, r in subset.iterrows()}

        # Insert an extra blank row at Sliding -> Fixed transitions
        if (previous_folder is not None
                and previous_base_type == "Sliding"
                and curr_base_type == "Fixed"):
            final_rows.append([])   # padded to max_len below

        # Insert an extra blank row between Fixed groups from different folders
        if (previous_folder is not None
                and previous_base_type == "Fixed"
                and curr_base_type == "Fixed"
                and previous_folder != curr_folder):
            final_rows.append([])   # padded to max_len below

        # Insert an extra blank row between Sliding groups from different folders
        if (previous_folder is not None
                and previous_base_type == "Sliding"
                and curr_base_type == "Sliding"
                and previous_folder != curr_folder):
            final_rows.append([])   # padded to max_len below

        # Build the data row: label followed by per-horizon metric groups
        row = [(file_type, 0)]
        for h in horizons_seen:
            if h in horizon_map:
                r = horizon_map[h]
                row.extend([
                    (r["MAE"],  r["MAE_CODE"]),
                    (r["RMSE"], r["RMSE_CODE"]),
                    (r["R2"],   r["R2_CODE"]),
                    r["LINES"],
                    r["COLUMNS"],
                    "",         # visual separator between horizons
                ])
            else:
                # Horizon absent for this file type: fill with empty cells
                row.extend([("", 0), ("", 0), ("", 0), "", "", ""])

        final_rows.append(row)

        previous_folder    = curr_folder
        previous_base_type = curr_base_type

    # Pad all rows to the same width so DataFrame construction does not error
    max_len = max(len(r) for r in final_rows)
    for r in final_rows:
        r.extend([("", 0)] * (max_len - len(r)))

    # ------------------------------------------------------------------
    # Build header row 2: column labels
    # ------------------------------------------------------------------
    # Row 1 is the file path (written by openpyxl after the data pass).
    # Row 2 has the column label header: "Type" then, for each horizon,
    # the four metric/shape labels plus a blank separator column.

    header_row = ["Type"]
    for h in horizons_seen:
        header_row.extend([f"MAE ({h})", f"RMSE ({h})", f"R² ({h})", "Lines", "Columns", ""])

    # Pad header to match data width
    while len(header_row) < max_len:
        header_row.append("")

    # ------------------------------------------------------------------
    # Write to Excel (values only, formatting applied separately)
    # ------------------------------------------------------------------
    # openpyxl cannot apply cell styles through pandas, so the approach
    # is: write raw values via pandas, reload the workbook with openpyxl,
    # then apply fonts, alignment and conditional fills cell by cell.

    clean_rows = [
        [cell[0] if isinstance(cell, tuple) else cell for cell in row]
        for row in final_rows
    ]

    # Reserve two header rows by prepending them before the data
    all_rows = [header_row] + clean_rows
    pd.DataFrame(all_rows).to_excel(output_excel, index=False, header=False)

    # ------------------------------------------------------------------
    # Excel formatting pass
    # ------------------------------------------------------------------

    wb = load_workbook(output_excel)
    ws = wb.active

    # --- Style definitions -------------------------------------------

    font_default = Font(name="Aptos Narrow", size=11)
    font_header  = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    font_title   = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left",   vertical="center")

    # Header background: dark blue-grey
    header_fill = PatternFill(start_color="2F4F6F", end_color="2F4F6F", fill_type="solid")
    # Title row background: slightly darker
    title_fill  = PatternFill(start_color="1C2E40", end_color="1C2E40", fill_type="solid")

    # Metric quality colour fills
    red_fill          = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_fill       = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

    total_cols = ws.max_column

    # ------------------------------------------------------------------
    # Row 1: source file path (merged title)
    # ------------------------------------------------------------------

    ws.insert_rows(1)   # push existing rows down by one to make room

    title_cell = ws.cell(row=1, column=1, value=f"Source : {txt_path}")
    title_cell.font      = font_title
    title_cell.fill      = title_fill
    title_cell.alignment = align_left

    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1,   end_column=total_cols
    )
    ws.row_dimensions[1].height = 20

    # ------------------------------------------------------------------
    # Row 2: column label header
    # ------------------------------------------------------------------

    for c_idx in range(1, total_cols + 1):
        cell = ws.cell(row=2, column=c_idx)
        cell.font      = font_header
        cell.fill      = header_fill
        cell.alignment = align_center
    ws.row_dimensions[2].height = 18

    # ------------------------------------------------------------------
    # Freeze rows 1 and 2 (pane anchored at A3)
    # ------------------------------------------------------------------

    ws.freeze_panes = "A3"

    # ------------------------------------------------------------------
    # Data rows: apply base font/alignment then quality-code colours
    # ------------------------------------------------------------------

    # Data starts at row 3 (row 1 = title, row 2 = header)
    for r_idx, row in enumerate(final_rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)

            if isinstance(val, tuple):
                text, code = val
                cell.value = text

                if code == -1:
                    cell.fill = red_fill
                elif code == 1:
                    cell.fill = light_orange_fill
                elif code == 2:
                    cell.fill = orange_fill
            else:
                cell.value = val

            cell.font      = font_default
            cell.alignment = align_center

    # ------------------------------------------------------------------
    # Auto-fit column widths
    # ------------------------------------------------------------------
    # openpyxl has no built-in auto-fit, so we iterate every cell and
    # track the maximum string length per column, then add a small
    # padding margin before setting column_dimensions.width.

    MIN_WIDTH = 8
    PADDING   = 3   # extra characters to avoid text being clipped

    col_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            # Merged cells span multiple columns; skip them for width calc
            if isinstance(cell, type(ws.cell(1, 1))) and cell.coordinate in ws.merged_cells:
                continue

            col_letter = get_column_letter(cell.column)
            text_len   = len(str(cell.value))
            col_widths[col_letter] = max(col_widths.get(col_letter, MIN_WIDTH),
                                         text_len + PADDING)

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_excel)
    print(f"Excel workbook written and formatted: {output_excel}")





def run(RESULT_PATH):

    print("\n" * 3)
    print("=" * 60)
    print("RESULT PARSER PIPELINE")
    print("=" * 60)

    input_txt    = RESULT_PATH + "\PROACT - RF Results.txt"
    output_excel = RESULT_PATH + "\PROACT - RF Results.xlsx"

    # Set convert_death_time=True when the log contains Death-target results
    # and MAE/RMSE should be expressed in days rather than interval counts.
    parse_results_to_excel(input_txt, output_excel, convert_death_time=False)

    # Open the generated workbook immediately after creation
    os.startfile(output_excel)